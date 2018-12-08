from django.http import HttpResponse, Http404,HttpResponseRedirect
from django.template import loader,Context,RequestContext
from symptomchecker.models import Diseases_Symptoms
from .forms import symform
from django.db.models import Q
from recommended_doctors.models import DoctorData
import pandas as pd
import numpy as np
import pyrenn as prn
from profile_user.models import user_details
from openpyxl import Workbook,load_workbook
list=[]
type=[]
def index(request):
    all_diseases = Diseases_Symptoms.objects.all()
    template=loader.get_template('symptomchecker/index.html')
    context = Context({
        'all_diseases' : all_diseases ,
    })
    return HttpResponse(template.render(context))

def detail(request, disease_id):
    disease12 = Diseases_Symptoms.objects.get(pk=disease_id)
    if disease12:
        template = loader.get_template('symptomchecker/detail.html')
        context = Context({
            'disease': disease12,
        })
        return HttpResponse(template.render(context))

    elif Diseases_Symptoms.DoesNotExist:
        raise Http404("NOT FOUND!")

def finddisease(request):
    del list[:]
    if request.method=='POST':
        q1=request.POST.get("symptom1")
        q2=request.POST.get("symptom2")
        q3 = request.POST.get("symptom3")
        q4 = request.POST.get("symptom4")
        q5 = request.POST.get("symptom5")
        q6 = request.POST.get("symptom6")
        q7 = request.POST.get("symptom7")
        q8 = request.POST.get("symptom8")
        q9 = request.POST.get("symptom9")
        q10 = request.POST.get("symptom10")
        q11= request.POST.get("symptom11")
        q12 = request.POST.get("symptom12")
        q13 = request.POST.get("symptom13")
        q14= request.POST.get("symptom14")
        q15 = request.POST.get("symptom15")
        q16= request.POST.get("symptom16")
        q17 = request.POST.get("symptom17")
        q18 = request.POST.get("symptom18")
        q19= request.POST.get("symptom19")
        q20 = request.POST.get("symptom20")
        q21= request.POST.get("symptom21")
        q22 = request.POST.get("symptom22")
        q23 = request.POST.get("symptom23")
        q24 = request.POST.get("symptom24")
        q25 = request.POST.get("symptom25")
        q26 = request.POST.get("symptom26")
        q27 = request.POST.get("symptom27")
        q28 = request.POST.get("symptom28")
        q29 = request.POST.get("symptom29")
        q30 = request.POST.get("symptom30")
        q31= request.POST.get("symptom31")
        q32 = request.POST.get("symptom32")
        q33 = request.POST.get("symptom33")
        q34= request.POST.get("symptom34")
        q35 = request.POST.get("symptom35")
        q36 = request.POST.get("symptom36")
        q37= request.POST.get("symptom37")
        q38= request.POST.get("symptom38")
        q39= request.POST.get("symptom39")
        q40 = request.POST.get("symptom40")
        q41 = request.POST.get("symptom41")
        q42 = request.POST.get("symptom42")
        q43 = request.POST.get("symptom43")
        q44 = request.POST.get("symptom44")
        q45= request.POST.get("symptom45")
        q46 = request.POST.get("symptom46")
        q47 = request.POST.get("symptom47")
        q48 = request.POST.get("symptom48")
        q49 = request.POST.get("symptom49")
        q50 = request.POST.get("symptom50")
        q = [q1,q2,q3,q4,q5,q6,q7,q8,q9,q10,q11,q12,q13,q14,q15,q16,q17,q18,q19,q20,q21,q22,q23,q24,q25,
             q26,q27,q28,q29,q30,q31,q32,q33,q34,q35,q36,q37,q38,q39,q40,q41,q42,q43,q44,q45,q46,q47,q48,q49,q50]
        s=[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,
           0,0,0]

        for i in range(len(q)):
            #for j in range(len(s)):
            if q[i]=='on':
                s[i]=1
        wb = load_workbook('test.xlsx',data_only=True)

        #wb = load_workbook('dataset.xlsx', data_only=True)
        ws = wb.active
        ws['CZ2'].value = s[0]
        ws['DA2'].value = s[1]
        ws['DB2'].value = s[2]
        ws['DC2'].value = s[3]
        ws['DD2'].value = s[4]
        ws['DE2'].value = s[5]
        ws['DF2'].value = s[6]
        ws['DG2'].value = s[7]
        ws['DH2'].value = s[8]
        ws['DI2'].value = s[9]
        ws['DJ2'].value = s[10]
        ws['DK2'].value = s[11]
        ws['DL2'].value = s[12]
        ws['DM2'].value = s[13]
        ws['DN2'].value = s[14]
        ws['DO2'].value = s[15]
        ws['DP2'].value = s[16]
        ws['DQ2'].value = s[17]
        ws['DR2'].value = s[18]
        ws['DS2'].value = s[19]
        ws['DT2'].value = s[20]
        ws['DU2'].value = s[21]
        ws['DV2'].value = s[22]
        ws['DW2'].value = s[23]
        ws['DX2'].value = s[24]
        ws['DY2'].value = s[25]
        ws['DZ2'].value = s[26]
        ws['EA2'].value = s[27]
        ws['EB2'].value = s[28]
        ws['EC2'].value = s[29]
        ws['ED2'].value = s[30]
        ws['EE2'].value = s[31]
        ws['EF2'].value = s[32]
        ws['EG2'].value = s[33]
        ws['EH2'].value = s[34]
        ws['EI2'].value = s[35]
        ws['EJ2'].value = s[36]
        ws['EK2'].value = s[37]
        ws['EL2'].value = s[38]
        ws['EM2'].value = s[39]
        ws['EN2'].value = s[40]
        ws['EO2'].value = s[41]
        ws['EP2'].value = s[42]
        ws['EQ2'].value = s[43]
        ws['ER2'].value = s[44]
        ws['ES2'].value = s[45]
        ws['ET2'].value = s[46]
        ws['EU2'].value = s[47]
        ws['EV2'].value = s[48]
        ws['EW2'].value = s[49]
        wb.save('test.xlsx')
        df = pd.ExcelFile('test.xlsx').parse('Sheet')
        P0 = np.array(
            [df['s1'].values, df['s2'].values, df['s3'].values, df['s4'].values, df['s5'].values, df['s6'].values,
             df['s7'].values, df['s8'].values, df['s9'].values, df['s10'].values, df['s11'].values, df['s12'].values,
             df['s13'].values, df['s14'].values, df['s15'].values, df['s16'].values, df['s17'].values, df['s18'].values,
             df['s19'].values, df['s20'].values, df['s21'].values, df['s22'].values, df['s23'].values, df['s24'].values,
             df['s25'].values, df['s26'].values, df['s27'].values, df['s28'].values, df['s29'].values, df['s30'].values,
             df['s31'].values, df['s32'].values, df['s33'].values, df['s34'].values, df['s35'].values, df['s36'].values,
             df['s37'].values, df['s38'].values, df['s39'].values, df['s40'].values, df['s41'].values, df['s42'].values,
             df['s43'].values, df['s44'].values, df['s45'].values, df['s46'].values, df['s47'].values, df['s48'].values,
             df['s49'].values, df['s50'].values])
        net = prn.loadNN('C:\Users\STUTI\Desktop\minor_final.csv')

        diseases = ['AIDS (acquired immuno-deficiency syndrome)', 'Adhesion', 'Affect labile', 'Alzheimers disease',
                    'Anemia', 'Aphasia', 'Asthma', 'Biliary calculus', 'Bipolar disorder', 'Carcinoma prostate',
                    'Cholecystitis', 'Chronic alcoholic intoxication', 'Chronic kidney failure',
                    'Chronic obstructive airway disease', 'Coronary arteriosclerosis', 'Decubitus ulcer',
                    'Degenerative polyarthritis', 'Deglutition disorder', 'Dehydration', 'Depressive disorder',
                    'Diverticulosis', 'Pulmonary Embolism', 'Encephalopathy', 'Endocarditis', 'Epilepsy',
                    'Heart Failure', 'Kidney Failure', 'Fibroid tumor', 'Gastritis', 'Gout', 'Hepatitis',
                    'Hernia hiatal', 'Hyperbilirubinemia', 'Hypercholesterolemia', 'Hyperglycemia', 'Hypothyroidism',
                    'Ileus', 'Incontinence', 'Infection urinary tract', 'Influenza', 'Insufficiency renal', 'Lymphoma',
                    'Malignant neoplasm of breast', 'Malignant neoplasm of prostate', 'Malignant tumor of colon',
                    'Myocardial infarction', 'Neoplasm', 'Neoplasm Metastasis', 'Obesity', 'Obesity Morbid']
        y = prn.NNOut(P0, net)
        z = y
        np.around(y, 0, z)
        for i in range(len(z)):
            if z[i][0] == 1:
                list.append(diseases[i])

        if request.session['first_name'] is None:
            template = loader.get_template('symptomchecker/index.html')
            context = RequestContext(request, {'checked': list, 'checked1': ws})
            return HttpResponse(template.render(context))
        else:
            template = loader.get_template('symptomchecker/indexloggedin.html')
            context = RequestContext(request, {'checked': list, 'checked1': ws, 'name': request.session['first_name']})
            return HttpResponse(template.render(context))


    else:
        if request.session['first_name'] is None:
            sform = symform()
            template = loader.get_template('symptomchecker/symtest.html')
            context = RequestContext(request, {'symform': sform})
            return HttpResponse(template.render(context))
        else:
            sform = symform()
            template = loader.get_template('symptomchecker/symptomc.html')
            context = RequestContext(request, {'symform': sform,'name': request.session['first_name']})
            return HttpResponse(template.render(context))


def recommenddoctor(request):
    disease_type={'AIDS (acquired immuno-deficiency syndrome)': 'Physician',
                  'Adhesion':'Radiology',
                  'Affect labile':'Psychiatry',
                 'Alzheimers disease':'Neurology',
                    'Anemia':'Hematology', 'Aphasia':'Neurology'
                     , 'Asthma':'Physician', 'Biliary calculus':'Gastroenteology',
                  'Bipolar disorder':'Psychiatry', 'Carcinoma prostate':'Urology & Kidney Transplant',
                    'Cholecystitis':'Gastroenteology', 'Chronic alcoholic intoxication':'Physician',
                  'Chronic kidney failure':'Nephrology',
                    'Chronic obstructive airway disease':'Paediatric Pulmonology', 'Coronary arteriosclerosis':'Cardiology',
                  'Decubitus ulcer':'Dermatology',
                    'Degenerative polyarthritis':'Orthopaedic', 'Deglutition disorder':'Physiopathology',
                  'Dehydration':'Physician', 'Depressive disorder':'Psychiatry',
                    'Diverticulosis':'Oncology', 'Pulmonary Embolism':'Paediatric Pulmonology', 'Encephalopathy':'Neuro Sciences',
                  'Endocarditis':'Cardiology', 'Epilepsy':'Neurology',
                    'Heart Failure':'Cardiology', 'Kidney Failure':'Nephrology', 'Fibroid tumor':'Obstetrician',
                  'Gastritis':'Gastroenteology', 'Gout':'Internal Medicine', 'Hepatitis':'Hepatology',
                    'Hernia hiatal':'General Surgery', 'Hyperbilirubinemia':'Physician',
                  'Hypercholesterolemia':'Gastroenteology', 'Hyperglycemia':'Physician', 'Hypothyroidism':'Endocrinoloy',
                    'Ileus':'Gastroenteology', 'Incontinence':'Urology & Kidney Transplant',
                  'Infection urinary tract':'Urology & Kidney Transplant', 'Influenza':'Physician',
                  'Insufficiency renal':'Nephrology', 'Lymphoma':'Physician',
                    'Malignant neoplasm of breast':'General Surgery', 'Malignant neoplasm of prostate':'Oncology',
                  'Malignant tumor of colon':'Oncology',
                    'Myocardial infarction':'Cardiology', 'Neoplasm':'General Surgery',
                  'Neoplasm Metastasis':'General Surgery', 'Obesity':'Physiotherapist',
                  'Obesity Morbid':'Physiotherapist'}
    for i in range(len(list)):
        for key,value in disease_type.iteritems():
            if list[i]==key:
                type.append(value)

    doctors=DoctorData.objects.filter(Q(type__exact=type[0])|
                                      Q(type__exact=type[1])
                                      )
    if request.session['first_name'] is None:
        template = loader.get_template('symptomchecker/test1.html')
        context = RequestContext(request, {'index': doctors})
        return HttpResponse(template.render(context))
    else:
        template = loader.get_template('symptomchecker/loggedintest.html')
        context = RequestContext(request, {'index': doctors,'name':request.session['first_name']})
        return HttpResponse(template.render(context))








